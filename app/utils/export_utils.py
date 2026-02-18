# -*- coding: utf-8 -*-
"""데이터 내보내기 유틸리티"""
from io import BytesIO
from datetime import datetime
from flask import send_file

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def create_excel_workbook(title="Report"):
    """
    Excel workbook 생성

    Args:
        title: 워크북 제목

    Returns:
        Workbook 객체
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl이 설치되어 있지 않습니다. pip install openpyxl을 실행하세요.")

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel 시트 이름은 최대 31자
    return wb, ws


def style_header_row(ws, row_num=1, column_count=1):
    """
    헤더 행 스타일링

    Args:
        ws: Worksheet 객체
        row_num: 헤더 행 번호
        column_count: 컬럼 수
    """
    header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    for col in range(1, column_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border


def style_data_rows(ws, start_row=2, end_row=None, column_count=1):
    """
    데이터 행 스타일링

    Args:
        ws: Worksheet 객체
        start_row: 시작 행
        end_row: 끝 행 (None이면 마지막 행까지)
        column_count: 컬럼 수
    """
    if end_row is None:
        end_row = ws.max_row

    data_font = Font(name='맑은 고딕', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center')
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for row in range(start_row, end_row + 1):
        # 짝수 행 배경색
        fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid') if row % 2 == 0 else None

        for col in range(1, column_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
            if fill:
                cell.fill = fill


def auto_adjust_column_width(ws, min_width=10, max_width=50):
    """
    컬럼 너비 자동 조정

    Args:
        ws: Worksheet 객체
        min_width: 최소 너비
        max_width: 최대 너비
    """
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def add_title_row(ws, title, column_count):
    """
    제목 행 추가

    Args:
        ws: Worksheet 객체
        title: 제목 텍스트
        column_count: 컬럼 수
    """
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = Font(name='맑은 고딕', size=14, bold=True, color='1F4E78')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 제목 행 높이 조정
    ws.row_dimensions[1].height = 30


def add_info_row(ws, info_text, row_num):
    """
    정보 행 추가 (날짜, 생성자 등)

    Args:
        ws: Worksheet 객체
        info_text: 정보 텍스트
        row_num: 행 번호
    """
    ws.insert_rows(row_num)
    info_cell = ws.cell(row=row_num, column=1)
    info_cell.value = info_text
    info_cell.font = Font(name='맑은 고딕', size=9, color='666666')
    info_cell.alignment = Alignment(horizontal='left', vertical='center')


def create_excel_response(wb, filename):
    """
    Excel 파일 다운로드 응답 생성

    Args:
        wb: Workbook 객체
        filename: 파일명 (확장자 제외)

    Returns:
        Flask send_file response
    """
    # BytesIO 객체에 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 파일명 생성 (날짜 포함)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    full_filename = f"{filename}_{timestamp}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=full_filename
    )


# ==================== 특정 데이터 내보내기 함수 ====================

def export_students_to_excel(students):
    """
    학생 목록을 Excel로 내보내기

    Args:
        students: Student 객체 리스트

    Returns:
        Flask response
    """
    wb, ws = create_excel_workbook("학생 목록")

    # 제목 추가
    add_title_row(ws, "전체 학생 목록", 7)
    add_info_row(ws, f"생성일시: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M')}", 2)

    # 헤더
    headers = ['번호', '이름', '학년', '등급', '이메일', '연락처', '생년월일']
    ws.append([])  # 빈 행
    ws.append(headers)

    # 헤더 스타일
    style_header_row(ws, row_num=4, column_count=len(headers))

    # 데이터
    for idx, student in enumerate(students, start=1):
        ws.append([
            idx,
            student.name,
            student.grade or '-',
            student.tier or '-',
            student.email or '-',
            student.phone or '-',
            student.birth_date.strftime('%Y-%m-%d') if student.birth_date else '-'
        ])

    # 스타일링
    style_data_rows(ws, start_row=5, column_count=len(headers))
    auto_adjust_column_width(ws)

    return create_excel_response(wb, "학생목록")


def export_courses_to_excel(courses):
    """
    수업 목록을 Excel로 내보내기

    Args:
        courses: Course 객체 리스트

    Returns:
        Flask response
    """
    wb, ws = create_excel_workbook("수업 목록")

    # 제목 추가
    add_title_row(ws, "전체 수업 목록", 8)
    add_info_row(ws, f"생성일시: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M')}", 2)

    # 헤더
    headers = ['번호', '수업명', '수업코드', '강사', '등급', '수강생', '세션', '상태']
    ws.append([])
    ws.append(headers)

    style_header_row(ws, row_num=4, column_count=len(headers))

    # 데이터
    for idx, course in enumerate(courses, start=1):
        ws.append([
            idx,
            course.course_name,
            course.course_code or '-',
            course.teacher.name if course.teacher else '미배정',
            course.tier or '-',
            f"{course.enrolled_count}/{course.max_students}",
            f"{course.completed_sessions}/{course.total_sessions}",
            '진행중' if course.status == 'active' else '완료' if course.status == 'completed' else '취소'
        ])

    style_data_rows(ws, start_row=5, column_count=len(headers))
    auto_adjust_column_width(ws)

    return create_excel_response(wb, "수업목록")


def export_attendance_to_excel(attendance_data, course_name=None):
    """
    출석 데이터를 Excel로 내보내기

    Args:
        attendance_data: 출석 데이터 리스트 (dict 형태)
        course_name: 수업명 (선택)

    Returns:
        Flask response
    """
    wb, ws = create_excel_workbook("출석부")

    # 제목
    title = f"{course_name} 출석부" if course_name else "전체 출석부"
    add_title_row(ws, title, 7)
    add_info_row(ws, f"생성일시: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M')}", 2)

    # 헤더
    headers = ['날짜', '수업명', '학생명', '출석상태', '지각시간', '비고', '체크시간']
    ws.append([])
    ws.append(headers)

    style_header_row(ws, row_num=4, column_count=len(headers))

    # 데이터
    for data in attendance_data:
        status_text = {
            'present': '출석',
            'late': '지각',
            'absent': '결석',
            'excused': '결석(사유)'
        }.get(data.get('status'), '-')

        ws.append([
            data.get('date', '-'),
            data.get('course_name', '-'),
            data.get('student_name', '-'),
            status_text,
            data.get('late_minutes', '-'),
            data.get('notes', '-'),
            data.get('checked_at', '-')
        ])

    style_data_rows(ws, start_row=5, column_count=len(headers))
    auto_adjust_column_width(ws)

    return create_excel_response(wb, "출석부")


def export_payments_to_excel(payments):
    """
    결제 내역을 Excel로 내보내기

    Args:
        payments: Payment 객체 리스트

    Returns:
        Flask response
    """
    wb, ws = create_excel_workbook("결제 내역")

    # 제목 추가
    add_title_row(ws, "결제 내역", 8)
    add_info_row(ws, f"생성일시: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M')}", 2)

    # 헤더
    headers = ['날짜', '학생명', '수업명', '금액', '회차', '결제방법', '상태', '비고']
    ws.append([])
    ws.append(headers)

    style_header_row(ws, row_num=4, column_count=len(headers))

    # 데이터
    total_amount = 0
    for payment in payments:
        ws.append([
            payment.created_at.strftime('%Y-%m-%d'),
            payment.student.name if payment.student else '-',
            payment.course.course_name if payment.course else '-',
            payment.amount,
            payment.sessions_covered or '-',
            payment.payment_method or '-',
            '완료' if payment.status == 'completed' else '대기' if payment.status == 'pending' else '취소',
            payment.notes or '-'
        ])

        if payment.status == 'completed':
            total_amount += payment.amount

    # 합계 행
    total_row = ws.max_row + 1
    ws.append(['', '', '합계', total_amount, '', '', '', ''])

    # 합계 행 스타일
    total_cell = ws.cell(row=total_row, column=3)
    total_cell.font = Font(name='맑은 고딕', size=11, bold=True)
    amount_cell = ws.cell(row=total_row, column=4)
    amount_cell.font = Font(name='맑은 고딕', size=11, bold=True, color='FF0000')

    style_data_rows(ws, start_row=5, end_row=total_row-1, column_count=len(headers))
    auto_adjust_column_width(ws)

    return create_excel_response(wb, "결제내역")


def export_student_report_to_excel(student, enrollments, essays, attendance_stats):
    """
    학생 종합 리포트를 Excel로 내보내기

    Args:
        student: Student 객체
        enrollments: CourseEnrollment 객체 리스트
        essays: Essay 객체 리스트
        attendance_stats: 출석 통계 dict

    Returns:
        Flask response
    """
    wb, ws = create_excel_workbook("학생 리포트")

    # 제목
    add_title_row(ws, f"{student.name} 학생 종합 리포트", 4)
    add_info_row(ws, f"생성일시: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M')}", 2)

    # 학생 기본 정보
    ws.append([])
    ws.append(['학생 정보', '', '', ''])
    ws.append(['이름', student.name, '학년', student.grade or '-'])
    ws.append(['이메일', student.email or '-', '등급', student.tier or '-'])
    ws.append(['연락처', student.phone or '-', '생년월일', student.birth_date.strftime('%Y-%m-%d') if student.birth_date else '-'])

    # 출석 통계
    ws.append([])
    ws.append(['출석 통계', '', '', ''])
    ws.append(['총 세션', attendance_stats.get('total_sessions', 0), '출석', attendance_stats.get('present', 0)])
    ws.append(['지각', attendance_stats.get('late', 0), '결석', attendance_stats.get('absent', 0)])
    ws.append(['출석률', f"{attendance_stats.get('attendance_rate', 0):.1f}%", '', ''])

    # 수강 중인 수업
    ws.append([])
    ws.append(['수강 중인 수업', '', '', ''])
    ws.append(['수업명', '강사', '출석률', '상태'])
    style_header_row(ws, row_num=ws.max_row, column_count=4)

    for enrollment in enrollments:
        ws.append([
            enrollment.course.course_name,
            enrollment.course.teacher.name if enrollment.course.teacher else '-',
            f"{enrollment.attendance_rate:.1f}%",
            '진행중' if enrollment.status == 'active' else '완료'
        ])

    # 첨삭 기록
    ws.append([])
    ws.append(['첨삭 기록', '', '', ''])
    ws.append(['제출일', '버전', '상태', '점수'])
    style_header_row(ws, row_num=ws.max_row, column_count=4)

    for essay in essays[:20]:  # 최근 20개
        ws.append([
            essay.created_at.strftime('%Y-%m-%d'),
            f"v{essay.current_version}",
            '완료' if essay.is_finalized else '진행중',
            '-'  # 점수는 EssayResult에서 가져와야 함
        ])

    auto_adjust_column_width(ws)

    return create_excel_response(wb, f"{student.name}_종합리포트")
