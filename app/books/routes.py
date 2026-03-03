# -*- coding: utf-8 -*-
"""도서 관리 라우트"""
import re
from flask import render_template, redirect, url_for, flash, request, jsonify
from flask_login import login_required, current_user
from app.books import books_bp
from app.books.forms import BookForm
from app.books.isbn_service import ISBNService
from app.models import db, Book, EssayBook


@books_bp.route('/')
@login_required
def index():
    """도서 목록"""
    # 검색 및 필터
    search = request.args.get('search', '').strip()
    category_filter = request.args.get('category', '').strip()
    sort_by = request.args.get('sort', 'recent')

    # 기본 쿼리
    query = Book.query

    # 검색 (제목, 저자, ISBN)
    if search:
        query = query.filter(
            db.or_(
                Book.title.contains(search),
                Book.author.contains(search),
                Book.isbn.contains(search)
            )
        )

    # 카테고리 필터
    if category_filter:
        query = query.filter_by(category=category_filter)

    # 정렬
    if sort_by == 'title':
        query = query.order_by(Book.title.asc())
    elif sort_by == 'author':
        query = query.order_by(Book.author.asc())
    elif sort_by == 'recent':
        query = query.order_by(Book.created_at.desc())
    else:
        query = query.order_by(Book.created_at.desc())

    books = query.all()

    return render_template('books/index.html',
                         books=books,
                         search=search,
                         category_filter=category_filter,
                         sort_by=sort_by)


@books_bp.route('/new', methods=['GET', 'POST'])
@login_required
def new():
    """도서 추가"""
    form = BookForm()

    if form.validate_on_submit():
        book = Book(
            user_id=current_user.user_id,
            title=form.title.data,
            author=form.author.data if form.author.data else None,
            publisher=form.publisher.data if form.publisher.data else None,
            isbn=form.isbn.data if form.isbn.data else None,
            publication_year=form.publication_year.data if form.publication_year.data else None,
            category=form.category.data if form.category.data else None,
            description=form.description.data if form.description.data else None,
            cover_image_url=form.cover_image_url.data if form.cover_image_url.data else None
        )

        db.session.add(book)
        db.session.commit()

        flash(f'"{book.title}" 도서가 추가되었습니다.', 'success')
        return redirect(url_for('books.detail', book_id=book.book_id))

    return render_template('books/form.html',
                         form=form,
                         title='새 도서 추가',
                         is_edit=False)


@books_bp.route('/<book_id>')
@login_required
def detail(book_id):
    """도서 상세"""
    book = Book.query.get_or_404(book_id)

    # 이 도서를 참고한 첨삭 목록
    essay_relations = EssayBook.query.filter_by(book_id=book_id).all()
    essays = [rel.essay for rel in essay_relations]

    return render_template('books/detail.html',
                         book=book,
                         essays=essays)


@books_bp.route('/<book_id>/edit', methods=['GET', 'POST'])
@login_required
def edit(book_id):
    """도서 수정"""
    book = Book.query.get_or_404(book_id)

    # 권한 확인 (본인 또는 관리자만)
    if book.user_id != current_user.user_id:
        flash('권한이 없습니다.', 'error')
        return redirect(url_for('books.detail', book_id=book_id))

    form = BookForm(obj=book)

    if form.validate_on_submit():
        book.title = form.title.data
        book.author = form.author.data if form.author.data else None
        book.publisher = form.publisher.data if form.publisher.data else None
        book.isbn = form.isbn.data if form.isbn.data else None
        book.publication_year = form.publication_year.data if form.publication_year.data else None
        book.category = form.category.data if form.category.data else None
        book.description = form.description.data if form.description.data else None
        book.cover_image_url = form.cover_image_url.data if form.cover_image_url.data else None

        db.session.commit()

        flash(f'"{book.title}" 정보가 수정되었습니다.', 'success')
        return redirect(url_for('books.detail', book_id=book.book_id))

    return render_template('books/form.html',
                         form=form,
                         title=f'{book.title} 수정',
                         is_edit=True,
                         book=book)


@books_bp.route('/<book_id>/delete', methods=['POST'])
@login_required
def delete(book_id):
    """도서 삭제"""
    book = Book.query.get_or_404(book_id)

    # 권한 확인
    if book.user_id != current_user.user_id:
        flash('권한이 없습니다.', 'error')
        return redirect(url_for('books.detail', book_id=book_id))

    book_title = book.title

    db.session.delete(book)
    db.session.commit()

    flash(f'"{book_title}" 도서가 삭제되었습니다.', 'info')
    return redirect(url_for('books.index'))


# API: ISBN 조회
@books_bp.route('/api/isbn-lookup', methods=['POST'])
@login_required
def isbn_lookup():
    """ISBN으로 도서 정보 조회"""
    data = request.get_json(silent=True) or {}
    isbn = data.get('isbn', '').strip()

    if not isbn:
        return jsonify({'success': False, 'message': 'ISBN을 입력하세요.'}), 400

    # ISBN 조회
    book_info = ISBNService.lookup_isbn(isbn)

    if not book_info:
        return jsonify({'success': False, 'message': 'ISBN으로 도서 정보를 찾을 수 없습니다.'}), 404

    return jsonify({
        'success': True,
        'book': book_info
    })


# ─────────────────────────────────────────────
# 일괄 ISBN 등록
# ─────────────────────────────────────────────

@books_bp.route('/bulk-import', methods=['GET'])
@login_required
def bulk_import():
    """일괄 ISBN 등록 페이지"""
    return render_template('books/bulk_import.html')


@books_bp.route('/bulk-import/lookup', methods=['POST'])
@login_required
def bulk_import_lookup():
    """파일에서 ISBN 추출 후 병렬 조회 (AJAX)"""
    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'success': False, 'message': '파일을 선택하세요.'}), 400

    filename = file.filename.lower()
    isbns = []

    try:
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            from openpyxl import load_workbook
            import io
            wb = load_workbook(io.BytesIO(file.read()), read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        val = re.sub(r'[\-\s]', '', str(cell).strip())
                        if re.fullmatch(r'\d{10}|\d{13}', val):
                            isbns.append(val)
        elif filename.endswith('.txt') or filename.endswith('.csv'):
            content = file.read().decode('utf-8-sig', errors='replace')
            for line in content.splitlines():
                val = re.sub(r'[\-\s]', '', line.strip())
                if re.fullmatch(r'\d{10}|\d{13}', val):
                    isbns.append(val)
        else:
            return jsonify({'success': False, 'message': '.xlsx, .txt, .csv 파일만 지원합니다.'}), 400
    except Exception as e:
        return jsonify({'success': False, 'message': f'파일 파싱 오류: {e}'}), 400

    # 중복 제거 (순서 유지)
    isbns = list(dict.fromkeys(isbns))

    if not isbns:
        return jsonify({'success': False, 'message': 'ISBN을 찾을 수 없습니다. (10자리 또는 13자리 숫자)'}), 400

    if len(isbns) > 50:
        return jsonify({'success': False, 'message': f'한번에 최대 50개까지 처리 가능합니다. (감지된 ISBN: {len(isbns)}개)'}), 400

    # 이미 등록된 ISBN
    existing_isbns = {b.isbn for b in Book.query.filter(Book.isbn.in_(isbns)).all()}

    # 병렬 조회
    from concurrent.futures import ThreadPoolExecutor, as_completed

    def _lookup(isbn):
        info = ISBNService.lookup_isbn(isbn)
        if info:
            info['already_exists'] = isbn in existing_isbns
            return {'isbn': isbn, 'found': True, 'data': info}
        return {'isbn': isbn, 'found': False, 'data': None}

    results = []
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {executor.submit(_lookup, isbn): isbn for isbn in isbns}
        for future in as_completed(futures):
            results.append(future.result())

    # 원래 순서 복원
    order_map = {isbn: i for i, isbn in enumerate(isbns)}
    results.sort(key=lambda x: order_map.get(x['isbn'], 999))

    return jsonify({'success': True, 'results': results, 'total': len(isbns)})


@books_bp.route('/bulk-save', methods=['POST'])
@login_required
def bulk_save():
    """선택한 도서 일괄 저장 (AJAX)"""
    books_data = request.get_json(silent=True) or []
    if not books_data:
        return jsonify({'success': False, 'message': '저장할 도서를 선택하세요.'}), 400

    saved = skipped = 0
    for item in books_data:
        isbn  = (item.get('isbn') or '').strip()
        title = (item.get('title') or '').strip()
        if not title:
            skipped += 1
            continue
        if isbn and Book.query.filter_by(isbn=isbn).first():
            skipped += 1
            continue
        book = Book(
            user_id=current_user.user_id,
            title=title,
            author=item.get('author') or None,
            publisher=item.get('publisher') or None,
            isbn=isbn or None,
            publication_year=item.get('publication_year') or None,
            description=item.get('description') or None,
            cover_image_url=item.get('cover_image_url') or None,
        )
        db.session.add(book)
        saved += 1

    db.session.commit()
    msg = f'{saved}권 등록 완료'
    if skipped:
        msg += f', {skipped}권 건너뜀 (중복 또는 제목 없음)'
    return jsonify({'success': True, 'message': msg, 'saved': saved, 'skipped': skipped})
