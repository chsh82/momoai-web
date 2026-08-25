# -*- coding: utf-8 -*-
"""
모모의 책장 연간 커리큘럼 엑셀 파싱 + DB 반영 공용 로직.
CLI 스크립트(import_curriculum.py)와 웹 업로드(app/curriculum/routes.py)가 함께 사용함.

엑셀 형식: "연간 전체 리스트" 시트에 연도/분기/학년/주차/기간/도서명/저자(역자)/출판사/휴강여부
9개 열이 있어야 함. 같은 (연도, 분기, 학년, 주차) 조합은 재실행 시 덮어씀(upsert)되므로
같은 파일을 수정해서 다시 올려도 안전함.

도서 매칭: 제목의 " (연장)" 표시를 뗀 뒤 books.title과 정확히 일치하는 기존 도서가 있으면
그 도서를 재사용(연결)하고, 없으면 새로 생성함(is_curriculum=True, grade_tags에 학년 코드 추가).
같은 제목의 기존 도서가 여러 건이면 저자가 일치하는 쪽 -> 이미 is_curriculum인 쪽 -> 가장
먼저 등록된 쪽 순으로 하나를 고름.
"""
import re
import json
import uuid
from datetime import datetime

import openpyxl

from app.models import db
from app.models.book import Book
from app.models.curriculum import CurriculumWeek
from app.utils.grades import GRADE_TO_LV

SHEET_NAME = '연간 전체 리스트'
REQUIRED_HEADERS = ['연도', '분기', '학년', '주차', '기간', '도서명', '저자(역자)', '출판사', '휴강여부']

GRADE_LABEL_TO_CODE = {
    '초등 1학년': '초1', '초등 2학년': '초2', '초등 3학년': '초3',
    '초등 4학년': '초4', '초등 5학년': '초5', '초등 6학년': '초6',
    '중학교 1학년': '중1', '중학교 2학년': '중2', '중학교 3학년': '중3',
    '고등학교 1학년': '고1', '고등학교 2학년': '고2', '고등학교 3학년': '고3',
}

_EXT_MARKER_RE = re.compile(r'\s*\(연장\)\s*$')


def _clean_title(title: str) -> str:
    return _EXT_MARKER_RE.sub('', (title or '').strip()).strip()


def _normalize_author(author: str) -> str:
    return re.sub(r'[\s\W_]+', '', author or '').lower()


def parse_workbook(source) -> list[dict]:
    """source: 파일 경로(str) 또는 파일 객체(업로드 스트림)"""
    wb = openpyxl.load_workbook(source, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f'"{SHEET_NAME}" 시트를 찾을 수 없습니다. (시트 목록: {wb.sheetnames})')
    ws = wb[SHEET_NAME]

    header = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if header[:len(REQUIRED_HEADERS)] != REQUIRED_HEADERS:
        raise ValueError(f'헤더가 예상과 다릅니다.\n기대: {REQUIRED_HEADERS}\n실제: {header}')

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        year, quarter, grade_label, week_number, date_range, title, author, publisher, holiday_flag = r[:9]
        if not year or not quarter or not grade_label or week_number is None:
            continue
        grade_code = GRADE_LABEL_TO_CODE.get(str(grade_label).strip())
        if not grade_code:
            raise ValueError(f'알 수 없는 학년 라벨: {grade_label!r}')
        rows.append({
            'year': int(year),
            'quarter': str(quarter).strip(),
            'grade': grade_code,
            'week_number': int(week_number),
            'date_range': str(date_range).strip() if date_range else None,
            'title': str(title).strip() if title else '',
            'author': str(author).strip() if author else None,
            'publisher': str(publisher).strip() if publisher else None,
            'is_holiday': str(holiday_flag).strip() == '휴강' if holiday_flag else False,
        })
    return rows


def _pick_book(candidates: list[Book], author: str | None) -> Book:
    if author:
        na = _normalize_author(author)
        for c in candidates:
            if na and _normalize_author(c.author) == na:
                return c
    for c in candidates:
        if c.is_curriculum:
            return c
    return min(candidates, key=lambda c: c.created_at or datetime.min)


def _find_or_create_book(title_index: dict, title: str, author: str | None, publisher: str | None,
                          grade_code: str, owner_user_id: str, stats: dict) -> Book:
    # Book.grade_tags는 학년 코드가 아니라 추천 레벨 코드(LV1~LV10)로 저장하는 컨벤션
    # (templates/library/admin/book_form.html 체크박스 값 참고) - 그대로 넣으면 안 됨.
    lv_tag = GRADE_TO_LV.get(grade_code)

    candidates = title_index.get(title)
    if candidates:
        book = _pick_book(candidates, author)
        changed = False
        if not book.is_curriculum:
            book.is_curriculum = True
            changed = True
        if not book.author and author:
            book.author = author
            changed = True
        if not book.publisher and publisher:
            book.publisher = publisher
            changed = True
        tags = json.loads(book.grade_tags) if book.grade_tags else []
        if lv_tag and lv_tag not in tags:
            tags.append(lv_tag)
            book.grade_tags = json.dumps(tags, ensure_ascii=False)
            changed = True
        if changed:
            stats['books_updated'] += 1
        return book

    book = Book(
        book_id=str(uuid.uuid4()),
        user_id=owner_user_id,
        title=title,
        author=author,
        publisher=publisher,
        is_curriculum=True,
        grade_tags=json.dumps([lv_tag] if lv_tag else [], ensure_ascii=False),
    )
    db.session.add(book)
    title_index[title] = [book]
    stats['books_created'] += 1
    return book


def apply_import(rows: list[dict], owner_user_id: str) -> dict:
    """rows를 세션에 add/수정만 하고 커밋은 호출한 쪽에서 함(dry-run 등을 위해 분리)."""
    all_books = Book.query.all()
    title_index: dict[str, list[Book]] = {}
    for b in all_books:
        title_index.setdefault(b.title.strip(), []).append(b)

    stats = {'weeks_added': 0, 'weeks_updated': 0, 'books_created': 0, 'books_updated': 0, 'holiday_weeks': 0}

    for r in rows:
        book = None
        if not r['is_holiday'] and r['title']:
            clean_title = _clean_title(r['title'])
            if clean_title:
                book = _find_or_create_book(title_index, clean_title, r['author'], r['publisher'],
                                             r['grade'], owner_user_id, stats)
        else:
            stats['holiday_weeks'] += 1

        existing = CurriculumWeek.query.filter_by(
            year=r['year'], quarter=r['quarter'], grade=r['grade'], week_number=r['week_number'],
        ).first()
        if existing:
            existing.date_range = r['date_range']
            existing.is_holiday = r['is_holiday']
            existing.note = r['title'] if r['is_holiday'] else None
            existing.book_id = book.book_id if book else None
            existing.updated_at = datetime.utcnow()
            stats['weeks_updated'] += 1
        else:
            db.session.add(CurriculumWeek(
                year=r['year'], quarter=r['quarter'], grade=r['grade'], week_number=r['week_number'],
                date_range=r['date_range'], is_holiday=r['is_holiday'],
                note=r['title'] if r['is_holiday'] else None,
                book_id=book.book_id if book else None,
            ))
            stats['weeks_added'] += 1

    stats['rows_parsed'] = len(rows)
    return stats
